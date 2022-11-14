import openpyxl

from Consts.consts import Consts
import logging
from Utilities.LogUtil import Logger
log = Logger(__name__, logging.INFO)

def getRowCount(path, sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_row


def getColCount(path, sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_column


def getCellData(sheetName, rowNum, colNum):

    path = Consts.Excel_DATA
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    data = sheet.cell(row=rowNum, column=colNum).value
    data = str(data)
    log.logger.info("[Read Excel] data cell (sheet: "+sheetName+", row:"+str(rowNum)+", colum:"+str(colNum)+") is: "+data)
    return data


def setCellData(path, sheetName, rowNum, colNum, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    sheet.cell(row=rowNum, column=colNum).value = data
    workbook.save(path)


#path = "../excel/testdata.xlsx"
#sheetName = "Deals"

#rows = getRowCount(path,sheetName)
#cols = getColCount(path,sheetName)

#print(rows,"---",cols)

#print(getCellData("Pizza",2,2))
#setCellData(path,sheetName,1,4,"DOB")